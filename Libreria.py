import openpyxl                       #Libreria para leer archivos de Excel
import numpy as np                    #Libreria para procesamiento Matematico
import matplotlib.pyplot as plt       #Libreria para Graficar

from sklearn import manifold          #Funcion para TSNE
from sklearn import preprocessing     #Funcion para TSNE
from sklearn.cluster import MeanShift #Funcion para especificar el tipo de 
                                      #Clustering




# Matriz con los colores de la libreria de "matplotlib"
colors = ['aliceblue','antiquewhite','aqua','aquamarine','azure','beige',
         'bisque','black','blanchedalmond','blue','blueviolet','brown',
         'burlywood','cadetblue','chartreuse','chocolate','coral',
         'cornflowerblue','cornsilk','crimson','cyan','darkblue','darkcyan',
         'darkgoldenrod','darkgray','darkgreen','darkkhaki','darkmagenta',
         'darkolivegreen','darkorange','darkorchid','darkred','darksalmon',
         'darkseagreen','darkslateblue','darkslategray','darkturquoise',
         'darkviolet','deeppink','deepskyblue','dimgray','dodgerblue',
         'firebrick','floralwhite','forestgreen','fuchsia','gainsboro',
         'ghostwhite','gold','goldenrod','gray','green','greenyellow',
         'honeydew','hotpink','indianred','indigo','ivory','khaki','lavender',
         'lavenderblush','lawngreen','lemonchiffon','lightblue','lightcoral',
         'lightcyan','lightgoldenrodyellow','lightgreen','lightgray',
         'lightpink','lightsalmon','lightseagreen','lightskyblue',
         'lightslategray','lightsteelblue','lightyellow','lime','limegreen',
         'linen','magenta','maroon','mediumaquamarine','mediumblue',
         'mediumorchid','mediumpurple','mediumseagreen','mediumslateblue',
         'mediumspringgreen','mediumturquoise','mediumvioletred','midnightblue',
         'mintcream','mistyrose','moccasin','navajowhite','navy','oldlace',
         'olive','olivedrab','orange','orangered','orchid','palegoldenrod',
         'palegreen','paleturquoise','palevioletred','papayawhip','peachpuff',
         'peru','pink','plum','powderblue','purple','red','rosybrown',
         'royalblue','saddlebrown','salmon','sandybrown','seagreen','seashell',
         'sienna','silver','skyblue','slateblue','slategray','snow',
         'springgreen','steelblue','tan','teal','thistle','tomato','turquoise',
         'violet','wheat','white','whitesmoke','yellow','yellowgreen']

"""
La funcion Tabla arma una matriz de nxm donde n es el numero de filas
especificados entre las variables Min y Max. La Variable m es el numero de 
columnas especificado entre Start y Stop. la ultima variable es la de criterio
para la clusterizacion.
Lee el valor de cada celda dentro del rango identificado, si el valor de la
celda no esta vacio, se guarda en su respectiva fila.
otro criterio de seleccion es que todos los elementos dentro de cada fila tengan
un valor, si una de las columnas no cuenta con un criterio satisfactorio, toda
la fila es descartada, puesto que no se podria realizar un analisis completo 
en ese grupo de datos.
por ultimo se crea un indice de las filas de las cuales se tomaron los datos
para su posterior uso.  
"""
def tabla (Min, Max, Start, Stop, Var): #Funcion para crear la tabla basada en 
                                        #en los datos del archivo de Excel
                                        #Variables:Primera Fila, Ultima Fila
                                        #Primera Columna, Ultima Columna,
                                        #Variable de Discriminacion
    a = 0       #Variable Utilizada para identificar celdas con los parametros
                #correctos
    n = []      #Matriz donde se almacenan las variables extraidas
    d = []      #Vector auxiliar para armar la Matriz
    
    c = np.zeros((1, Stop-Start))   #Crear el primer valor de un vector de 
                                    #dimensiones 1x(cantidad de variables)
    c = np.delete(c, 0, 0)          #Eliminar la primera posicion del vector,
                                    #pero mantiene la estructura
    
    for i in range (Min, Max, 1):   #Desde la primera Columna hasta la ultima
        a = 0                       #Bandera en 0
        for j in range (Start, Stop, 1): #desde la primera Fila hasta la ultima
            if sheet.cell(row=i, column=j).value is None or sheet.cell(row=i, column=Var).value is None:
                        #Si la celda analizada no contiene un valor y 
                        #la celda de la variable tampoco                
                a = 1   #Bandera en 1
            else:       #Sino
                d.append(sheet.cell(row=i, column=j).value)  
                        #Se guarda el calor de la celda en "d"
        if a == 0:      #si la bandera no fue levantada
            c = np.vstack((c,d))    #se carga el valor del vector en otro 
                                    #creando una Matriz
            n.append(i) #crea un vector indice de la tabla original del Excel
            
        d = []          #Vaciar el vector "d" para volver a llenarlo
        
    return (c, n)       #Devuelve la matriz y el indice
"""
La funcion Armar junta dos matrices de la misma dimension nxm creando una
nueva matriz nxm con los elementos de las originales.
esto se realiza debido a que el proceso de TSNE asigna una dimension
diferente a dos matrices de distintas dimensiones, las cuales no se pueden
sobreponer, esto quiere decir que si aplico el desdoblamiento TSNE a la matriz
original y a la de Clusters, no puedo graficarlas en un mismo eje cartesiano.
"""
def armar (Vector, Cluster):        #Funcion para armar una nueva matriz 
                                    #uniendo otras  dos
    Cluster = np.vstack((Cluster, Vector))  #Vector Cluster es la suma de 
                                            #los dos importados
    return Cluster                          #Devuelve la nueva Matriz

"""
La funcion Desarmar requiere la matriz combinada y el punto de division. 
el punto de division es igual al numero de elementos introducidos, es decir el 
numero de clusters.
crea una nueva matriz con los valores desdoblados de los clusters y elimina
estos valores de la matriz importada, dejandola unicamente con los valores de 
los elementos originales desdoblados   
"""
def desarmar (Vector, ClusterNum):   #Funcion para desarmar una matriz en un 
                                     #punto de Division
    a = Vector.shape[1]              #Una constante de la dimension de las 
                                     #columnas del vector
    b = []                           #Vector auxiliar 
    NewCluster = np.zeros((1, a))    #Crear el primer valor de un vector de 
                                     #dimensiones 1xa
    NewCluster = np.delete(NewCluster, 0, 0) #Eliminar la primera posicion del 
                                             #vector mantieniendo la estructura  
    for i in range(0, ClusterNum, 1):   #Desde 0 hasta el punto de division
        b.append(Vector[i])             #Crea un nuevo vector con los valores
                                        #de la fila correspondiente
        NewCluster = np.vstack((NewCluster, b)) #Armar una Matriz con los 
                                                #vectores obtenidos
        b = []                          #Vaciar el Vector "b"
        
    for i in range(0, ClusterNum, 1):   #Desde 0 hasta el punto de division
        Vector = np.delete(Vector, 0, 0)#Borrar la primera fila de la Matriz
        
    return (Vector, NewCluster)     #Devuelve el vector reducido y un nuevo 
                                    #vector con lo restado
"""
La funcion Var crea una matriz de 1xn donde n es el numero de filas validas
utilizadas para crear la tabla principal. los elementos de esta matriz son
los de la variable de criterio para la asignacion de colores
"""
def var(n, m):                      #Funcion para identificar de la Variable
    aux = []                        #Vector Auxiliar
    for i in range(0, len(n), 1):   #Desde 0 hasta el ultimovector de n
        aux.append(sheet.cell(row=n[i], column=m).value)  #Cargar el valor de 
                                                          #la variable criterio
    return aux                      #Devolver el vector de la variable
    
"""
La funcion TaC_Range crea un vector con la lista de elementos unicos en la 
lista de clusters y los divide dentro de intervalos creados en la variable aux1.
"""
def TaC_Range(Indice):  #Funcion para identificar Rangos de elementos identicos
                        #en la lista 
    Ind, aux = np.unique(Indice, return_counts=True)    #Ver cuantos elementos 
                                                        #unicos estan en la 
                                                        #lista
    aux1 = np.linspace(100, 0, num=21)    #Variable con intervalos de 5 en 5
    New=np.zeros((len(Indice),1))         #Cantidad de silas igual a la 
                                          #Cantidad de elementos
    for i in range(0,len(aux1),1):        #De 0 hasta la cantidad de intervalos
        for j in range(0,len(Indice),1):  #De 0 hasta la cantidad de elemtneos
            if Indice[j]<=aux1[i]:        #Si el elemento es menor al rango
                New[j]=i                  #Elemento guardado en la matriz
    return New                            #Retornar matriz


"""
La funcion TaC crea un vector con la lista de elementos unicos en la lista de 
los clusters.
"""
def TaC(Indice):        #Funcion para identificar Rangos de elementos identicos
                        #en la lista
    Ind, aux = np.unique(Indice, return_counts=True)    #Ver cuantos elementos 
                                                        #unicos estan en la 
                                                        #lista
    for i in range(0,len(Ind),1):        #De 0 hasta la cantidad de intervalos
        for j in range(0,len(Indice),1): #De 0 hasta la cantidad de elemtneos
            if Indice[j] == Ind[i]:      #Si el elemento es igual al del indice
                Indice[j] = i            #Elemento guardado en la matriz
    return Indice                        #Retornar matriz

"""
La funcion Procesamiento es la funcion principal, dentro de la cual se
contempla la llamada con las variables correspondientes al rango de datos
dentro de la tabla de excel y la variable de discriminacion para la graficacion
Primeramente se crea la tabla con los parametros definidos, se escala cada
elemento de la tabla entre 0-10, posterior a esto se lo normaliza.
Se aplica Clusterizacion , guarda el cluster al que cada elemento de la lista 
pertenece en un documento de texto externo para facilitar su analisis manual
de ser necesario.
Fusiona las matrices de clusters y de datos para aplicar un desdoblamiento TSNE
luego se separa la matriz en dos con la misma cantidad de filas original, pero
con los elementos reducidos a 2 dimensiones.
por ultimo se analiza la variable de discrimizacion para graficar cada elemento
en un eje de coordenadas y marcar los centros de los clusters identificados.
"""
def Procesamiento (Min, Max, Start, Stop, Var): #Funcion principal, Variables:
                                                #Variables:Primera Fila,
                                                #Ultima Fila, Primera Columna,
                                                #Ultima Columna,
                                                #Variable de discriminacion
    """
    Funcion
    """
    x = Var             #x asume el valor de la variable
    
    Tabla, N = tabla(Min, Max, Start, Stop, x)  #Llama a la funcion tabla y 
                                                #almacena la Matriz en Tabla 
                                                #y la variable en N
    
    min_max_scaler = preprocessing.MinMaxScaler()
    #asignacion de nombre a la funcion de escalameinto.
    X_train_minmax = min_max_scaler.fit_transform(Tabla)*10
    #Escala los cada valor de la tabla entre 0 y 1 y multiplica cada uno para 
    #llevar cada uno a una escala entre 0-10
    Tabla = preprocessing.normalize(X_train_minmax, norm='l2')
    #Normaliza los datos de la Tabla
    
    """
    Clustering
    """
    print 'Clustering'  #Imprimir al comienzo de la clusterizacion
    ms = MeanShift()    #Tipo de Clusterizacion
    n_components = 2    #Parametros minimos para Clusterizacion
    n_neighbors = 10    #Parametros minimos para Clusterizacion
    ms.fit(Tabla)       #Asignar formato necesario a al tabla 
    labels = ms.labels_ #Guardar los clusters identificados
    cluster_centers = ms.cluster_centers_   #Coordenadas de los centros de 
                                            #cada Cluster
    n_clusters_ = len(np.unique(labels))    #Cantidad de Clusters encontrados
    print("Number of estimated clusters:", n_clusters_) #Mostrar el numero de 
                                                        #Clusters
    
    """
    Guardar los Clusters
    """
    f = open('file.txt', 'wb')      #Abrir un archivo de texto llamado "File"
    for i in range(len(labels)):    #desde 0 hasta la cantidad de Elementos
        f.write("%i \n" % labels[i])    #Guardar a que Cluster pertenece cada 
                                        #elemento en una nueva linea
    f.close()                       #Cierra el documento
    
    """
    TSNE
    """
    print 'TSNE'        #Imprimir al comienzo del desdoblamiento por TSNE
    Combinadas = armar(Tabla, cluster_centers)  #Armar la matris combinando la 
                                                #original con los clusters 
                                                #encontrados
    tsne = manifold.TSNE(n_components=n_components, init='pca', random_state=0)
    #Parametros de configuracion para el TSNE
    Desdoblada = tsne.fit_transform(Combinadas) #Desdoblamiento de la matriz
    
    g, h = desarmar(Desdoblada, n_clusters_)    #Desarmar la matriz: en g los
                                                #datos originales y en h las
                                                #coordenadas de los centros de
                                                #los clusters
    
    
    if Var == 2:                #Si la variable es la edad
        print Var               #Imprimir la variable
        Var = var(N, x)         #recolectar los valores de las variables
        Var = TaC_Range(Var)    #Dividir en rangos y asignar colores
        plt.scatter(g[:, 0], g[:, 1], c=Var)        #Graficar puntos
    else:
        Var = var(N, x)         #recolectar los valores de las variables
        Var = TaC(Var)          #designar los colores de cada Cluster.
        plt.scatter(g[:, 0], g[:, 1], c=Var)        #Graficar puntos
        
    for i in range(len(h)):     #De 0 hasta la cantidad de los Clusters
        plt.plot(h[i][0], h[i][1], c='r', marker='x', markersize = 10) 
        #Graficar los centros de los Clusters
    plt.axis('tight')   #Parametro de Grafica
    plt.show()          #Mostrar las Graficas
